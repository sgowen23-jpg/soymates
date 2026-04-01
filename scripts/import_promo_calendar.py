#!/usr/bin/env python3
"""
Import Promo Calendar from Excel into Supabase promo_calendar table.

Excel: C:/Users/sgowe/Downloads/1. UPDATED - 03032026 Promo Calendar_VAP_By MSO_updated 11.02.26 (2).xlsx
Table: promo_calendar

Each retailer sheet has a different layout. SHEET_CONFIG defines the exact
row/column positions for each sheet, verified by inspection of the actual file.
"""

import openpyxl
import urllib.request
import urllib.parse
import urllib.error
import json
import datetime
import sys

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_PATH = (
    r"C:\Users\sgowe\Downloads"
    r"\1. UPDATED - 03032026 Promo Calendar_VAP_By MSO_updated 11.02.26 (2).xlsx"
)

SUPABASE_URL = "https://zrqnnyugnxtlmkboegyn.supabase.co"
SUPABASE_KEY = "sb_publishable_pFWQYGJhjM-BGPNbvwGUQg_dK0izygj"
TABLE = "promo_calendar"

# Sheet name → retailer label (handle trailing-space variants)
SHEET_TO_RETAILER = {
    "IGA Promotions":    "IGA",
    "Ritchies Promotions": "Ritchies",
    "Foodworks.":        "Foodworks",
    "Foodland":          "Foodland",
    "Drakes  ":          "Drakes",
    "Drakes ":           "Drakes",
    "Drakes":            "Drakes",
}

# Per-sheet layout config (all indices are 0-based)
#   header_row  : row index containing column labels + Customer Week date cells
#   data_start  : row index of the first data row
#   type_col    : column index of the "Type" cell ("1. Price" / "2. Case Deal")
#   desc_col    : column index of the product description
#   sku_col     : column index of the SKU (None if not present)
#   dates_start : first column index that may contain date cells in the header row
SHEET_CONFIG = {
    "IGA":       {"header_row": 3, "data_start": 4,  "type_col": 3, "desc_col": 5, "sku_col": None, "dates_start": 9},
    "Ritchies":  {"header_row": 5, "data_start": 6,  "type_col": 1, "desc_col": 3, "sku_col": 2,    "dates_start": 5},
    "Foodworks": {"header_row": 1, "data_start": 2,  "type_col": 3, "desc_col": 6, "sku_col": 5,    "dates_start": 10},
    "Foodland":  {"header_row": 5, "data_start": 6,  "type_col": 1, "desc_col": 3, "sku_col": 2,    "dates_start": 6},
    "Drakes":    {"header_row": 5, "data_start": 6,  "type_col": 3, "desc_col": 6, "sku_col": 5,    "dates_start": 10},
}

BATCH_SIZE = 200

# ── Helpers ───────────────────────────────────────────────────────────────────

def format_display_value(val):
    """Return a display string for a numeric promo value."""
    if isinstance(val, (int, float)):
        if val > 0:
            return f"${val:.2f}".rstrip("0").rstrip(".")
        return str(val)
    return str(val)


def parse_value(raw):
    """
    Parse a promo cell value.
    Returns (value, display_value, skip) where:
      - skip=True  → ignore this cell entirely (no promo)
      - value      → float or None
      - display_value → string label shown in the UI
    """
    if raw is None:
        return None, None, True

    # datetime objects are header/metadata cells, not promo values
    if isinstance(raw, datetime.datetime):
        return None, None, True

    if isinstance(raw, (int, float)):
        v = float(raw)
        return v, format_display_value(v), False

    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return None, None, True

        # Long text → annotation, skip
        if len(stripped) > 30:
            return None, None, True

        # Known non-value strings
        if stripped.lower() in ("deleted", "n/a", "tbc", ".", "#ref!"):
            return None, None, True

        # Multi-buy patterns like "2 for $5"
        lower = stripped.lower()
        if "for $" in lower or " for " in lower:
            return None, stripped, False

        # Try to coerce numeric strings ("2", "10.88", "$2.80")
        try:
            v = float(stripped.replace("$", "").strip())
            return v, format_display_value(v), False
        except ValueError:
            pass

        # Short non-numeric string — treat as display label
        return None, stripped, False

    return None, None, True


def parse_promo_type(raw):
    """Map raw Type cell to 'price' or 'case_deal', or None to skip."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "1. Price":
        return "price"
    if s == "2. Case Deal":
        return "case_deal"
    return None


# ── Core parser ───────────────────────────────────────────────────────────────

def parse_sheet(ws, retailer):
    """
    Parse a worksheet using its per-sheet config.
    Returns (records, skipped_count).
    """
    cfg = SHEET_CONFIG[retailer]
    rows_data = list(ws.iter_rows(values_only=True))

    header_row  = rows_data[cfg["header_row"]]
    data_start  = cfg["data_start"]
    type_col    = cfg["type_col"]
    desc_col    = cfg["desc_col"]
    sku_col     = cfg.get("sku_col")
    dates_start = cfg["dates_start"]

    # Build week_cols: col_index → "YYYY-MM-DD" for every date cell in the header row
    week_cols = {}
    for i in range(dates_start, len(header_row)):
        cell = header_row[i]
        if isinstance(cell, datetime.datetime):
            week_cols[i] = cell.strftime("%Y-%m-%d")

    if not week_cols:
        print(f"  WARNING: No date columns found in {retailer} header row {cfg['header_row']} — skipping sheet.")
        return [], 0

    records = []
    skipped = 0

    # sort_order: preserves original Excel row ordering per product.
    # Both price and case_deal rows for the same product share the same value.
    product_sort_order = {}
    next_sort_order = 1

    for row in rows_data[data_start:]:
        # Type column determines price vs case_deal
        raw_type   = row[type_col] if type_col < len(row) else None
        promo_type = parse_promo_type(raw_type)
        if promo_type is None:
            skipped += 1
            continue

        # Description
        description = None
        if desc_col < len(row) and row[desc_col] is not None:
            description = str(row[desc_col]).strip() or None

        # SKU
        sku = None
        if sku_col is not None and sku_col < len(row) and row[sku_col] is not None:
            v = row[sku_col]
            if isinstance(v, float) and v == int(v):
                sku = str(int(v))
            else:
                sku = str(v).strip() or None

        # Assign sort_order on first appearance of each (description, sku) pair
        product_key = (description, sku)
        if product_key not in product_sort_order:
            product_sort_order[product_key] = next_sort_order
            next_sort_order += 1
        sort_order = product_sort_order[product_key]

        # Iterate date columns
        for col_idx, week_start in week_cols.items():
            raw_val = row[col_idx] if col_idx < len(row) else None
            value, display_value, skip = parse_value(raw_val)

            if skip:
                continue

            records.append({
                "retailer":          retailer,
                "product_description": description,
                "sku":               sku,
                "week_start":        week_start,
                "promo_type":        promo_type,
                "value":             value,
                "display_value":     display_value,
                "sort_order":        sort_order,
            })

    return records, skipped


# ── Supabase HTTP helpers ─────────────────────────────────────────────────────

def supabase_delete(retailer):
    """Delete all existing rows for this retailer."""
    url = f"{SUPABASE_URL}/rest/v1/{TABLE}?retailer=eq.{urllib.parse.quote(retailer)}"
    req = urllib.request.Request(
        url,
        method="DELETE",
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status
    except urllib.error.HTTPError as e:
        print(f"  ERROR deleting {retailer}: {e.code} {e.read().decode()}")
        return e.code


def supabase_insert_batch(records):
    """POST a batch of records to Supabase."""
    url = f"{SUPABASE_URL}/rest/v1/{TABLE}"
    body = json.dumps(records).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, None
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode()


# ── Main ──────────────────────────────────────────────────────────────────────

def run(dry_run=False, dry_run_limit=3):
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    sheets = [s for s in wb.sheetnames if s in SHEET_TO_RETAILER]
    if not sheets:
        print("ERROR: No matching sheets found. Check SHEET_TO_RETAILER mapping.")
        sys.exit(1)

    print(f"Found sheets: {sheets}\n")

    for sheet_name in sheets:
        retailer = SHEET_TO_RETAILER[sheet_name]
        ws = wb[sheet_name]

        print(f"--- {retailer} ({sheet_name}) ---")
        records, skipped = parse_sheet(ws, retailer)

        # Count unique products
        unique = len({r["product_description"] for r in records})
        weeks  = len({r["week_start"] for r in records})
        print(f"  Parsed  : {len(records)} promo entries ({unique} products, {weeks} weeks)")
        print(f"  Skipped : {skipped} rows (no valid type)")

        if dry_run:
            print(f"  DRY-RUN — first {dry_run_limit} records:")
            for i, r in enumerate(records[:dry_run_limit]):
                print(f"  [{i+1}] {json.dumps(r, default=str)}")
            print()
            continue

        # Live run: delete existing data then insert fresh
        print(f"  Deleting existing {retailer} rows from Supabase…")
        status = supabase_delete(retailer)
        print(f"  Delete status: {status}")

        inserted = 0
        for i in range(0, len(records), BATCH_SIZE):
            batch  = records[i : i + BATCH_SIZE]
            status, err = supabase_insert_batch(batch)
            if err:
                print(f"  ERROR on batch {i // BATCH_SIZE + 1}: {status} — {err}")
            else:
                inserted += len(batch)
                print(f"  Inserted batch {i // BATCH_SIZE + 1}: {len(batch)} rows (total {inserted})")

        print(f"  Done — {inserted} rows inserted for {retailer}.\n")


if __name__ == "__main__":
    live = "--live" in sys.argv

    if live:
        print("=" * 60)
        print("LIVE MODE — writing to Supabase")
        print("=" * 60)
        run(dry_run=False)
    else:
        print("=" * 60)
        print("DRY-RUN MODE — showing first 3 records per sheet")
        print("=" * 60)
        run(dry_run=True, dry_run_limit=3)
        print()
        print("To import all retailers into Supabase, run:")
        print("  python import_promo_calendar.py --live")
